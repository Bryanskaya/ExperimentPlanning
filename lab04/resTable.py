from openpyxl import Workbook, load_workbook
from os import startfile


class ExcelTable(object):
    filename = 'results.xlsx'
    error_msg = 'Ошибка: не удалось открыть файл'
    header = ['x0', 'x1', 'x2', 'x3', 'x4',
              'x1x2', 'x1x3', 'x1x4',
              'x2x3', 'x2x4',
              'x3x4',
              'x1x2x3', 'x1x2x4', 'x1x3x4',
              'x2x3x4',
              'x1x2x3x4',
              'x1^2 - a', 'x2^2 - a', 'x3^2 - a', 'x4^2 - a',
              'y', 'y нел',
              '|y - y нел|']

    def __init__(self, filename):
        self.filename = filename

    def create(self, table):
        book = Workbook()
        tbl = book.active

        tbl.append(self.header)

        numRows = len(table)

        for i in range(numRows):
            lenTable = len(table[i])
            stringTable = []
            for j in range(1, lenTable + 1):
                if type(table[i][j - 1]) is int:
                #if j < lenTable - 4:
                    stringTable.append(table[i][j - 1])
                else:
                    stringTable.append('%.3f' % table[i][j - 1])
            tbl.append(stringTable)

        try:
            book.save(self.filename)
        except PermissionError:
            print(self.error_msg)

    def add_one_row(self, row):
        book = load_workbook(self.filename)
        tbl = book.active

        numRows = len(row)
        for i in range(16, numRows):
            row[i] = '%.3f' % row[i]
        tbl.append(row)

        try:
            book.save(self.filename)
        except PermissionError:
            print(self.error_msg)

    def open(self):
        startfile(self.filename)


class PlanTable(object):
    def __init__(self, filename):
        self.filename = filename
        self.resultExcel = ExcelTable(self.filename)

    def show(self, table):
        self.resultExcel.create(table)
        self.resultExcel.open()

    def addRow(self, row):
        self.resultExcel.add_one_row(row)
        self.resultExcel.open()
