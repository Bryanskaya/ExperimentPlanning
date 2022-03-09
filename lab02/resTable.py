from openpyxl import Workbook, load_workbook
from os import startfile


class ExcelTable(object):
    filename = 'results.xlsx'
    error_msg = 'Ошибка'
    header = ['x0', 'x1', 'x2', 'x1x2',
              'y', 'y л', 'y нел',
              '|y - y л|', '|y - y нел|']

    def __init__(self, filename):
        self.filename = filename

    def create(self, table):
        book = Workbook()
        tbl = book.active

        tbl.append(self.header)

        numRows = len(table)

        for i in range(numRows):
            lenTable = len(table[i])
            stringTable = []
            for j in range(1, lenTable + 1):
                if j < lenTable - 4:
                    stringTable.append(table[i][j - 1])
                else:
                    stringTable.append('%.3f' % table[i][j - 1])
            tbl.append(stringTable)

        try:
            book.save(self.filename)
        except PermissionError:
            print(self.error_msg)

    def add_one_row(self, row):
        book = load_workbook(self.filename)
        tbl = book.active

        numRows = len(row)
        for i in range(4, numRows):
            row[i] = '%.4f' % row[i]
        tbl.append(row)

        try:
            book.save(self.filename)
        except PermissionError:
            print(self.error_msg)

    def open(self):
        startfile(self.filename)